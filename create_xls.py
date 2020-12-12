#!/usr/bin/python3.8

import openpyxl, xlrd, xlwt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.cell import Cell
import os
from os import listdir
import sys
import csv
import fnmatch
import datetime


###### DEFINE ######
DIR = "/home/mriantf/script_skripsi"
DT_DIR = DIR+"/DATA_REPORT"
SCRIPT = DIR+"/script"
SRC_IMG= DIR+"/OUTPUT"
OUTPUT_PDF = DIR+"/OUTPUT_PDF"
GTYPE = sys.argv[1]

templatexls = SCRIPT+"/excel_tmp.xlsx"
book = openpyxl.load_workbook(templatexls)
ws = book.active
# all_sheet = book.sheetnames

csv_list = listdir(DT_DIR+'/'+GTYPE)

for filecsv in csv_list:
    with open (DT_DIR+'/'+GTYPE+'/'+filecsv) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')
        #line_count = 0
        for row in csv_reader:
            IDREPORT = row[0]
            EMAIL = row[2]
            TITLE = row[3]
            RRDTITLE = row[6]
            RRDTITLE2 = RRDTITLE.replace(" ", "_")
            RRDTITLE3 = RRDTITLE2.replace("/","-")
            PERIODIC = row[7]
            filelist = fnmatch.filter(os.listdir(SRC_IMG+'/'+GTYPE), "*"+RRDTITLE3+"*")
            count_array = len(filelist)+7
            idx = 1
            for imglist in filelist:
                img = Image(SRC_IMG+'/'+GTYPE+'/'+imglist)
                for i in range(7, count_array):
                    cellg = i+1
                    ws['A'+i] = idx + ". Traffic Pemakaian " + RRDTITLE3
                    
                # for col in range(7, count_array):
                #     ws['A'+col] = imglist
                #     ws.add_image(img, 'B2')

                #ws['A1'] = imglist
                #ws.add_image(img, 'B2')
                idx += 1
    book.save(OUTPUT_PDF+'/'+GTYPE+'/'+"ReportID"+IDREPORT+"_"+TITLE+".xlsx")


